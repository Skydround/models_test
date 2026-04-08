#!/usr/bin/env python3
"""
Step 3: Evaluate model responses and generate comparison report.

This script:
1. Loads questions and model responses from the database
2. Uses an OpenRouter model as an LLM judge
3. Generates an Excel comparison spreadsheet
"""

import argparse
import os
import sys
import html
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from tqdm import tqdm
import json

sys.path.append(str(Path(__file__).parent))
from utils import QuestionDatabase, excel_safe_sheet_name, exam_sort_key, parse_exam_name

load_dotenv()

EVALUATOR_MODEL = os.getenv('OPENROUTER_EVALUATOR_MODEL', 'openai/gpt-4o-mini')

openrouter_client = OpenAI(
    base_url="https://openrouter.ai/api/v1",
    api_key=os.getenv('OPENROUTER_API_KEY'),
    default_headers={
        "HTTP-Referer": "https://github.com/models_test",
        "X-Title": "Models Test - Response Evaluation"
    }
)


def normalize_optional_text(value: str | None) -> str | None:
    """Convert placeholder strings to missing values."""
    if value is None:
        return None
    normalized = str(value).strip()
    if not normalized or normalized.lower() in {'null', 'none', 'brak'}:
        return None
    return normalized


def normalize_score(raw_score: float, max_points: int | None) -> float:
    """Map model scores to the scoring granularity implied by max_points."""
    score = max(0.0, min(1.0, float(raw_score)))
    if not max_points or max_points <= 1:
        return 1.0 if score >= 0.999 else 0.0

    step = 1 / max_points
    normalized = round(score / step) * step
    return round(max(0.0, min(1.0, normalized)), 2)


def normalize_evaluation_result(result: dict, max_points: int | None) -> dict:
    """Post-process evaluator output to match the exam scoring scheme."""
    normalized = dict(result)
    score = normalize_score(normalized.get('score', 0.0), max_points)
    normalized['score'] = score
    normalized['is_correct'] = score >= 0.999
    return normalized


def sanitize_excel_value(value):
    """Remove characters that openpyxl cannot write into worksheet cells."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub('', value)
    return value


def sanitize_excel_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Return a DataFrame copy with Excel-safe cell values."""
    return df.map(sanitize_excel_value)


def evaluate_response_with_llm(question: dict, response: str, correct_answer: str) -> dict:
    """
    Use an OpenRouter model as judge to evaluate a matura response.
    """

    answer_explanation = normalize_optional_text(question.get('answer_explanation')) or 'Not provided'
    context_text = normalize_optional_text(question.get('context_text')) or 'Not provided'
    exam_name = question.get('exam_name') or 'Unknown exam'
    question_type = question.get('question_type') or 'unknown'
    max_points = question.get('max_points')
    max_points_text = max_points if max_points is not None else 'Not provided'

    prompt = f"""You are an expert evaluator for Polish matura exam responses across all subjects.

Evaluate the answer in the context of the full exam task, not only by string matching.

EXAM:
{exam_name}

QUESTION TYPE:
{question_type}

MAX POINTS:
{max_points_text}

QUESTION:
{question['question_text']}

QUESTION CONTEXT:
{context_text}

CORRECT ANSWER:
{correct_answer if correct_answer else 'Not provided'}

ANSWER EXPLANATION / MARK SCHEME:
{answer_explanation}

STUDENT RESPONSE:
{response}

TASK:
Evaluate whether the student's response deserves full credit.

SCORING RULES:
1. For multiple choice or true/false tasks: exact correctness is required.
2. For short-answer tasks: accept semantic equivalence, equivalent notation, and minor wording differences.
3. For extended or open-ended tasks: require the essential points, claims, calculations, interpretations, or conclusions expected by the mark scheme.
4. For subjects other than Polish, prioritize subject accuracy over style.
5. Ignore minor grammar, spelling, or formatting issues unless they change the meaning.
6. If max_points = 1, score must be binary: only 0.0 or 1.0.
7. If max_points > 1, score must reflect point fractions achievable in the task, i.e. multiples of 1/max_points.
8. Examples: for 2 points use 0.0, 0.5, 1.0; for 3 points use 0.0, 0.33, 0.67, 1.0; for 4 points use 0.0, 0.25, 0.5, 0.75, 1.0.

Respond in JSON format:
{{
  "score": 0.0 to 1.0,
  "is_correct": true/false,
  "notes": "Brief explanation of your evaluation"
}}

Return ONLY the JSON, no other text."""

    try:
        completion = openrouter_client.chat.completions.create(
            model=EVALUATOR_MODEL,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0
        )
        
        result_text = completion.choices[0].message.content
        result = json.loads(result_text)
        usage = getattr(completion, 'usage', None)
        if usage is not None:
            result['prompt_tokens'] = getattr(usage, 'prompt_tokens', 0) or 0
            result['completion_tokens'] = getattr(usage, 'completion_tokens', 0) or 0
        return normalize_evaluation_result(result, max_points)
        
    except Exception as e:
        print(f"\n⚠️  Evaluation error: {e}")
        return {
            'score': 0.0,
            'is_correct': False,
            'notes': f'Evaluation failed: {str(e)}'
        }


def generate_excel_report(db: QuestionDatabase, output_path: str):
    """Generate Excel comparison spreadsheet"""
    
    print("\n📊 Generating Excel report...")
    
    # Get all data from database
    conn = db.conn
    
    # Query to get all questions with responses
    query = """
    SELECT 
        q.id as question_id,
        q.exam_name,
        q.question_number,
        q.question_text,
        q.question_type,
        q.max_points,
        q.correct_answer,
        mr.model_name,
        mr.response,
        mr.latency_ms,
        mr.tokens_used,
        mr.cost_usd,
        e.score,
        e.is_correct,
        e.evaluator_notes
    FROM questions q
    LEFT JOIN model_responses mr ON q.id = mr.question_id
    LEFT JOIN evaluations e ON mr.id = e.response_id
    ORDER BY q.exam_name, q.question_number, mr.model_name
    """
    
    df = pd.read_sql_query(query, conn)
    
    if df.empty:
        print("❌ No data to export")
        return

    exam_metadata = df['exam_name'].apply(parse_exam_name).apply(pd.Series)
    df = pd.concat([df, exam_metadata], axis=1)
    df.rename(columns={'name': 'exam_code'}, inplace=True)
    
    # Create pivot table for easier comparison
    pivot_data = []
    
    for question_id in df['question_id'].unique():
        q_data = df[df['question_id'] == question_id]
        
        row = {
            'Exam': q_data.iloc[0]['exam_name'],
            'Exam Label': q_data.iloc[0]['session_label'],
            'Subject': q_data.iloc[0]['subject_label'],
            'Year': q_data.iloc[0]['year'],
            'Level': q_data.iloc[0]['level'],
            'Question #': q_data.iloc[0]['question_number'],
            'Question': q_data.iloc[0]['question_text'][:100] + '...',  # Truncate
            'Type': q_data.iloc[0]['question_type'],
            'Max Points': q_data.iloc[0]['max_points'],
            'Correct Answer': q_data.iloc[0]['correct_answer']
        }
        
        # Add each model's response and score
        for _, resp in q_data.iterrows():
            if pd.notna(resp['model_name']):
                model = resp['model_name']
                row[f'{model}_Response'] = resp['response']
                row[f'{model}_Score'] = resp['score']
                row[f'{model}_Correct'] = '✓' if resp['is_correct'] else '✗'
                row[f'{model}_Latency_ms'] = resp['latency_ms']
                row[f'{model}_Cost_USD'] = resp['cost_usd']
        
        pivot_data.append(row)
    
    pivot_df = pd.DataFrame(pivot_data)
    pivot_df_excel = sanitize_excel_dataframe(pivot_df)
    
    models = sorted(df['model_name'].dropna().unique())
    exams  = sorted(df['exam_name'].dropna().unique(), key=exam_sort_key)

    # ── helpers ────────────────────────────────────────────────────────────
    def pct(series):
        v = series.fillna(0).astype(float).mean()
        return f"{v * 100:.1f}%"

    def build_model_summary(subset, label_col=None, label_val=None) -> list[dict]:
        rows = []
        for model in models:
            md = subset[subset['model_name'] == model]
            if md.empty:
                continue
            row = {}
            if label_col:
                row[label_col] = label_val
            row.update({
                'Model': model,
                'Questions': len(md),
                'Correct': int(md['is_correct'].fillna(0).sum()),
                'Accuracy %': pct(md['is_correct']),
                'Avg Score': round(md['score'].fillna(0).mean(), 3),
                'Avg Latency ms': round(md['latency_ms'].fillna(0).mean()),
                'Total Cost $': round(md['cost_usd'].fillna(0).sum(), 5),
                'Cost/Q $': round(md['cost_usd'].fillna(0).mean(), 6),
            })
            rows.append(row)
        return rows

    def build_type_breakdown(subset) -> list[dict]:
        rows = []
        for qtype in sorted(subset['question_type'].dropna().unique()):
            td = subset[subset['question_type'] == qtype]
            for model in models:
                md = td[td['model_name'] == model]
                if md.empty:
                    continue
                rows.append({
                    'Question Type': qtype,
                    'Model': model,
                    'Questions': len(md),
                    'Correct': int(md['is_correct'].fillna(0).sum()),
                    'Accuracy %': pct(md['is_correct']),
                    'Avg Score': round(md['score'].fillna(0).mean(), 3),
                })
        return rows

    def autofit_sheet(ws):
        for col_cells in ws.columns:
            length = max((len(str(c.value or '')) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 4, 60)

    def style_header_row(ws, fill_hex='C6EFCE'):
        fill = PatternFill('solid', fgColor=fill_hex)
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = bold
            cell.alignment = Alignment(horizontal='center')

    # ── write workbook ──────────────────────────────────────────────────────
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # ── 1. Summary (per model, overall) ─────────────────────────────────
        summary_df = sanitize_excel_dataframe(pd.DataFrame(build_model_summary(df)))
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        ws = writer.sheets['Summary']
        style_header_row(ws, 'BDD7EE')
        autofit_sheet(ws)

        # ── 2. By Exam (per exam × model) ───────────────────────────────────
        by_exam_rows = []
        for exam_name in exams:
            ed = df[df['exam_name'] == exam_name]
            meta = ed.iloc[0]
            rows = build_model_summary(
                ed, label_col='Exam', label_val=meta['session_label']
            )
            for r in rows:
                r['Subject'] = meta['subject_label']
                r['Year']    = meta['year']
                r['Level']   = meta['level']
                r['Exam Code'] = exam_name
            by_exam_rows.extend(rows)
        by_exam_df = sanitize_excel_dataframe(pd.DataFrame(by_exam_rows))
        # reorder columns
        front = ['Exam', 'Subject', 'Year', 'Level', 'Model']
        rest  = [c for c in by_exam_df.columns if c not in front + ['Exam Code']]
        by_exam_df = by_exam_df[front + rest + ['Exam Code']]
        by_exam_df.to_excel(writer, sheet_name='By Exam', index=False)
        ws = writer.sheets['By Exam']
        style_header_row(ws, 'FCE4D6')
        autofit_sheet(ws)

        # ── 3. By Question Type ─────────────────────────────────────────────
        type_df = sanitize_excel_dataframe(pd.DataFrame(build_type_breakdown(df)))
        type_df.to_excel(writer, sheet_name='By Type', index=False)
        ws = writer.sheets['By Type']
        style_header_row(ws, 'E2EFDA')
        autofit_sheet(ws)

        # ── 4. Per-exam detail sheets ────────────────────────────────────────
        for exam_name in exams:
            exam_pivot = pivot_df_excel[pivot_df_excel['Exam'] == exam_name]
            sheet_name = excel_safe_sheet_name(exam_name, fallback='Exam')
            exam_pivot.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            style_header_row(ws, 'FFF2CC')
            autofit_sheet(ws)

        # ── 5. Full Comparison (all exams, all models) ───────────────────────
        pivot_df_excel.to_excel(writer, sheet_name='All Questions', index=False)
        ws = writer.sheets['All Questions']
        style_header_row(ws, 'EDEDED')
        autofit_sheet(ws)

        # ── 6. Raw Data ──────────────────────────────────────────────────────
        sanitize_excel_dataframe(df).to_excel(writer, sheet_name='Raw Data', index=False)
        style_header_row(writer.sheets['Raw Data'], 'EDEDED')

    print(f"✅ Excel report saved to: {output_path}")


def generate_html_report(db: QuestionDatabase, output_path: str):
        """Generate a readable HTML comparison report."""

        print("\n🌐 Generating HTML report...")

        query = """
        SELECT
                q.id as question_id,
                q.exam_name,
                q.question_number,
                q.question_text,
                q.question_type,
                q.max_points,
                q.correct_answer,
                q.answer_explanation,
                q.context_text,
                mr.model_name,
                mr.response,
                mr.latency_ms,
                mr.tokens_used,
                mr.cost_usd,
                e.score,
                e.is_correct,
                e.evaluator_notes
        FROM questions q
        LEFT JOIN model_responses mr ON q.id = mr.question_id
        LEFT JOIN evaluations e ON mr.id = e.response_id
        ORDER BY q.exam_name, q.page_number, q.question_number, mr.model_name
        """

        df = pd.read_sql_query(query, db.conn)

        if df.empty:
                print("❌ No data to export to HTML")
                return

        exam_metadata = df['exam_name'].apply(parse_exam_name).apply(pd.Series)
        df = pd.concat([df, exam_metadata], axis=1)
        df.rename(columns={'name': 'exam_code'}, inplace=True)

        def as_text(value, fallback='Brak'):
            if pd.isna(value) or value is None:
                return fallback
            text = str(value).strip()
            if text == '' or text.lower() in {'null', 'none', 'brak'}:
                return fallback
            return text

        def as_html_block(value, fallback='Brak'):
                return html.escape(as_text(value, fallback)).replace('\n', '<br>')

        all_models = sorted(df['model_name'].dropna().unique())
        all_exams  = sorted(df['exam_name'].dropna().unique(), key=exam_sort_key)

        def pct_val(series):
            v = series.fillna(0).astype(float)
            if v.empty:
                return '—'
            return f'{v.mean() * 100:.1f}%'

        summary_rows = []
        for model in all_models:
                model_data = df[df['model_name'] == model]
                accuracy = model_data['is_correct'].fillna(0).astype(float).mean() * 100
                avg_score = model_data['score'].fillna(0).mean()
                correct   = int(model_data['is_correct'].fillna(0).sum())
                summary_rows.append(
                        f'<tr>'
                        f'<td>{html.escape(model)}</td>'
                        f'<td>{len(model_data)}</td>'
                        f'<td>{correct}</td>'
                        f'<td>{accuracy:.1f}%</td>'
                        f'<td>{avg_score:.3f}</td>'
                        f'<td>${model_data["cost_usd"].fillna(0).sum():.4f}</td>'
                        f'</tr>'
                )

        exam_summary_rows = []  # kept for compat but replaced by exam_stats_rows_html below

        question_sections_by_exam = {}
        for question_id in df['question_id'].unique():
                q_data = df[df['question_id'] == question_id]
                q_row = q_data.iloc[0]

                response_cards = []
                for _, resp in q_data.iterrows():
                        if pd.isna(resp['model_name']):
                                continue

                        score = 0.0 if pd.isna(resp['score']) else float(resp['score'])
                        if score >= 0.99:
                                score_class = 'score-high'
                        elif score >= 0.5:
                                score_class = 'score-mid'
                        else:
                                score_class = 'score-low'

                        is_correct = bool(resp['is_correct']) if not pd.isna(resp['is_correct']) else False
                        verdict = 'Correct' if is_correct else 'Needs review'

                        response_cards.append(
                                f"""
                                <article class="response-card">
                                    <div class="response-header">
                                        <h3>{html.escape(str(resp['model_name']))}</h3>
                                        <div class="badges">
                                            <span class="badge {score_class}">Score: {score:.2f}</span>
                                            <span class="badge {'badge-ok' if is_correct else 'badge-warn'}">{verdict}</span>
                                        </div>
                                    </div>
                                    <div class="meta">Latency: {0 if pd.isna(resp['latency_ms']) else int(resp['latency_ms'])} ms | Tokens: {0 if pd.isna(resp['tokens_used']) else int(resp['tokens_used'])} | Cost: ${0 if pd.isna(resp['cost_usd']) else float(resp['cost_usd']):.4f}</div>
                                    <div class="panel">
                                        <div class="panel-title">Model response</div>
                                        <div class="panel-body">{as_html_block(resp['response'])}</div>
                                    </div>
                                    <div class="panel">
                                        <div class="panel-title">Evaluator notes</div>
                                        <div class="panel-body">{as_html_block(resp['evaluator_notes'])}</div>
                                    </div>
                                </article>
                                """
                        )

                context_value = as_text(q_row['context_text'])
                context_block = ''
                if context_value != 'Brak':
                        context_block = f"""
                            <div class="panel context-block">
                                <div class="panel-title">Question context</div>
                                <div class="panel-body">{as_html_block(q_row['context_text'])}</div>
                            </div>
                        """

                question_sections_by_exam.setdefault(q_row['exam_name'], []).append(
                        f"""
                        <section class="question-card">
                            <div class="question-topline">
                                <span class="exam-pill">{html.escape(as_text(q_row['exam_name']))}</span>
                                <span class="session-pill">{html.escape(as_text(q_row['session_label']))}</span>
                                <span class="question-pill">Question {html.escape(as_text(q_row['question_number']))}</span>
                                <span class="type-pill">{html.escape(as_text(q_row['question_type']))}</span>
                                <span class="points-pill">{html.escape(as_text(q_row['max_points']))} pts</span>
                            </div>
                            <h2>{html.escape(as_text(q_row['question_text']))}</h2>
                            <div class="details-grid">
                                <div class="panel">
                                    <div class="panel-title">Correct answer</div>
                                    <div class="panel-body">{as_html_block(q_row['correct_answer'])}</div>
                                </div>
                                <div class="panel">
                                    <div class="panel-title">Mark scheme / explanation</div>
                                    <div class="panel-body">{as_html_block(q_row['answer_explanation'])}</div>
                                </div>
                            </div>
                            {context_block}
                            <div class="responses-grid">
                                {''.join(response_cards)}
                            </div>
                        </section>
                        """
                )

        grouped_sections = []
        subject_names = sorted(df['subject'].dropna().unique())
        for subject in subject_names:
                subject_rows = df[df['subject'] == subject]
                subject_label = subject_rows.iloc[0]['subject_label']
                exam_sections = []
                for exam_name in sorted(subject_rows['exam_name'].dropna().unique(), key=exam_sort_key):
                        exam_rows = subject_rows[subject_rows['exam_name'] == exam_name]
                        exam_sections.append(
                                f"""
                                <section class="exam-section" id="exam-{html.escape(exam_name)}">
                                    <div class="exam-header">
                                        <h3>{html.escape(exam_rows.iloc[0]['session_label'])}</h3>
                                        <div class="exam-code">{html.escape(exam_name)}</div>
                                    </div>
                                    {''.join(question_sections_by_exam.get(exam_name, []))}
                                </section>
                                """
                        )

                grouped_sections.append(
                        f"""
                        <section class="subject-section">
                            <div class="subject-header">
                                <h2>{html.escape(subject_label)}</h2>
                            </div>
                            {''.join(exam_sections)}
                        </section>
                        """
                )

        # ── helpers ──────────────────────────────────────────────────────────
        def bar_html(pct_str, width=None):
            """Return a progress-bar cell for a percentage string like '73.4%' or '—'."""
            if pct_str == '—':
                return f'<td class="heat-na">—</td>'
            val = float(pct_str.rstrip('%'))
            col = 'bar-hi' if val >= 65 else ('bar-mid' if val >= 40 else 'bar-lo')
            heat = 'heat-hi' if val >= 65 else ('heat-mid' if val >= 40 else 'heat-lo')
            w = round(val)
            return (
                f'<td class="bar-cell {heat}">'
                f'<div class="bar-wrap">'
                f'<div class="bar-bg"><div class="bar-fill {col}" style="width:{w}%"></div></div>'
                f'<span class="bar-pct">{pct_str}</span>'
                f'</div></td>'
            )

        def scorecard_bar(pct_str):
            if pct_str == '—':
                return ''
            val = float(pct_str.rstrip('%'))
            col = 'bar-hi' if val >= 65 else ('bar-mid' if val >= 40 else 'bar-lo')
            return (
                f'<div class="scorecard-bar-bg">'
                f'<div class="scorecard-bar-fill {col}" style="width:{round(val)}%"></div>'
                f'</div>'
            )

        # ── ranking table (sorted by accuracy) ──────────────────────────────
        model_acc = []
        for model in all_models:
            md = df[df['model_name'] == model]
            acc = md['is_correct'].fillna(0).astype(float).mean() * 100
            correct = int(md['is_correct'].fillna(0).sum())
            avg_score = md['score'].fillna(0).mean()
            cost = md['cost_usd'].fillna(0).sum()
            questions = len(md)
            model_acc.append((acc, model, correct, avg_score, cost, questions))
        model_acc.sort(reverse=True)

        ranking_rows = []
        for i, (acc, model, correct, avg_score, cost, questions) in enumerate(model_acc, 1):
            rank_cls = f'rank rank-{i}' if i <= 3 else 'rank'
            pct = f'{acc:.1f}%'
            col = 'bar-hi' if acc >= 65 else ('bar-mid' if acc >= 40 else 'bar-lo')
            ranking_rows.append(
                f'<tr>'
                f'<td><span class="{rank_cls}">{i}</span></td>'
                f'<td>{html.escape(model)}</td>'
                f'<td>{questions}</td>'
                f'<td>{correct}</td>'
                f'<td class="bar-cell"><div class="bar-wrap">'
                f'<div class="bar-bg"><div class="bar-fill {col}" style="width:{round(acc)}%"></div></div>'
                f'<span class="bar-pct">{pct}</span></div></td>'
                f'<td>{avg_score:.3f}</td>'
                f'<td>${cost:.4f}</td>'
                f'</tr>'
            )

        # sorted model list (by rank) for consistent column order
        ranked_models = [m for _, m, *_ in model_acc]
        model_header_cells = ''.join(f'<th>{html.escape(m)}</th>' for m in ranked_models)

        # ── per-type stats ───────────────────────────────────────────────────
        type_stats_rows = []
        for qtype in sorted(df['question_type'].dropna().unique()):
            td = df[df['question_type'] == qtype]
            cells = ''.join(
                bar_html(pct_val(td[td['model_name'] == m]['is_correct']))
                for m in ranked_models
            )
            type_stats_rows.append(f'<tr><td>{html.escape(qtype)}</td>{cells}</tr>')

        # ── per-exam stats (heatmap) ─────────────────────────────────────────
        exam_stats_rows_html = []
        for exam_name in all_exams:
            ed = df[df['exam_name'] == exam_name]
            label = html.escape(ed.iloc[0]['session_label'])
            anchor = f'exam-{exam_name}'
            q_count = ed['question_id'].nunique()
            cells = ''.join(
                bar_html(pct_val(ed[ed['model_name'] == m]['is_correct']))
                for m in ranked_models
            )
            exam_stats_rows_html.append(
                f'<tr><td><a href="#{anchor}">{label}</a><br><small style="color:var(--muted)">{q_count} pyt.</small></td>{cells}</tr>'
            )

        # ── per-exam scorecards ──────────────────────────────────────────────
        scorecard_items = []
        for exam_name in all_exams:
            ed = df[df['exam_name'] == exam_name]
            label = html.escape(ed.iloc[0]['session_label'])
            q_count = ed['question_id'].nunique()
            rows_html = ''
            for model in ranked_models:
                p = pct_val(ed[ed['model_name'] == model]['is_correct'])
                if p == '—':
                    continue
                rows_html += (
                    f'<div class="scorecard-row">'
                    f'<span class="scorecard-model" title="{html.escape(model)}">{html.escape(model)}</span>'
                    f'{scorecard_bar(p)}'
                    f'<span class="scorecard-pct">{p}</span>'
                    f'</div>'
                )
            scorecard_items.append(
                f'<div class="exam-scorecard">'
                f'<div class="exam-scorecard-header">'
                f'<a class="exam-scorecard-title" href="#exam-{html.escape(exam_name)}">{label}</a>'
                f'<span class="exam-scorecard-meta">{q_count} pytań</span>'
                f'</div>'
                f'{rows_html}'
                f'</div>'
            )
        scorecards_html = ''.join(scorecard_items)

        # ── nav links ────────────────────────────────────────────────────────
        nav_links = ''
        for subject in subject_names:
            sr = df[df['subject'] == subject]
            subject_label_safe = html.escape(sr.iloc[0]['subject_label'])
            exam_anchors = ''.join(
                f'<a class="nav-exam" href="#exam-{en}">{html.escape(df[df["exam_name"]==en].iloc[0]["session_label"])}</a>'
                for en in sorted(sr['exam_name'].dropna().unique(), key=exam_sort_key)
            )
            nav_links += f'<div class="nav-group"><span class="nav-subject">{subject_label_safe}</span>{exam_anchors}</div>'

        html_output = f"""<!DOCTYPE html>
<html lang="pl">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Models Test Comparison</title>
    <style>
        :root {{
            --bg: #f5efe4;
            --surface: #fffdf8;
            --surface-strong: #fff8ee;
            --text: #1f1a17;
            --muted: #6f6257;
            --border: #dfcfbb;
            --accent: #b85c38;
            --accent-soft: #f3d9bf;
            --ok: #2f7d4a;
            --warn: #9d5c0d;
            --bad: #a63f3f;
            --shadow: 0 14px 40px rgba(71, 48, 26, 0.08);
            --nav-h: 52px;
        }}
        * {{ box-sizing: border-box; }}
        body {{ margin: 0; font-family: Georgia, 'Times New Roman', serif; background: radial-gradient(circle at top, #fff7ea 0%, var(--bg) 45%, #efe4d3 100%); color: var(--text); }}
        /* sticky nav */
        .topnav {{ position: sticky; top: 0; z-index: 100; background: rgba(31,26,23,0.96); backdrop-filter: blur(8px); display: flex; gap: 0; overflow-x: auto; padding: 0 16px; height: var(--nav-h); align-items: center; scrollbar-width: thin; }}
        .nav-group {{ display: flex; align-items: center; gap: 6px; padding: 0 10px; border-right: 1px solid rgba(255,255,255,.12); }}
        .nav-group:last-child {{ border-right: none; }}
        .nav-subject {{ color: rgba(255,255,255,.45); font-size: .78rem; letter-spacing: .06em; text-transform: uppercase; white-space: nowrap; margin-right: 4px; }}
        .nav-exam {{ color: #f3d9bf; font-size: .88rem; text-decoration: none; white-space: nowrap; padding: 4px 8px; border-radius: 8px; transition: background .15s; }}
        .nav-exam:hover {{ background: rgba(243,217,191,.2); }}
        .page {{ max-width: 1500px; margin: 0 auto; padding: 32px 20px 60px; }}
        .hero {{ background: linear-gradient(135deg, rgba(184,92,56,.12), rgba(243,217,191,.75)); border: 1px solid var(--border); border-radius: 28px; padding: 28px; box-shadow: var(--shadow); }}
        h1 {{ margin: 0 0 8px; font-size: clamp(2rem, 4vw, 3.5rem); }}
        .subtitle {{ margin: 0; color: var(--muted); font-size: 1.05rem; }}
        .stats-wrap {{ display: flex; flex-direction: column; gap: 20px; margin-top: 24px; }}
        .stats-row {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(380px, 1fr)); gap: 20px; }}
        .table-box {{ background: var(--surface); border-radius: 18px; box-shadow: var(--shadow); overflow: hidden; }}
        .table-scroll {{ overflow-x: auto; }}
        .table-box caption {{ display: block; text-align: left; font-size: .9rem; font-weight: 700; letter-spacing: .05em; text-transform: uppercase; color: var(--muted); padding: 14px 18px 6px; }}
        .summary-table {{ width: 100%; border-collapse: collapse; min-width: 320px; }}
        .summary-table th, .summary-table td {{ padding: 10px 14px; border-bottom: 1px solid #efe2d3; text-align: left; font-size: .88rem; white-space: nowrap; }}
        .summary-table th {{ background: var(--surface-strong); }}
        .summary-table a {{ color: var(--accent); text-decoration: none; }}
        .summary-table a:hover {{ text-decoration: underline; }}
        .summary-table tbody tr:last-child td {{ border-bottom: none; }}
        .summary-table tbody tr:hover td {{ background: rgba(184,92,56,.04); }}
        .rank {{ display: inline-block; width: 22px; height: 22px; line-height: 22px; text-align: center; border-radius: 50%; font-size: .78rem; font-weight: 700; background: var(--accent-soft); color: var(--accent); }}
        .rank-1 {{ background: #ffd700; color: #7a5800; }}
        .rank-2 {{ background: #e0e0e0; color: #555; }}
        .rank-3 {{ background: #e8b97a; color: #6b3d00; }}
        .bar-cell {{ min-width: 140px; }}
        .bar-wrap {{ display: flex; align-items: center; gap: 8px; }}
        .bar-bg {{ flex: 1; height: 8px; background: #efe2d3; border-radius: 4px; overflow: hidden; min-width: 60px; }}
        .bar-fill {{ height: 100%; border-radius: 4px; }}
        .bar-hi {{ background: var(--ok); }}
        .bar-mid {{ background: var(--warn); }}
        .bar-lo {{ background: var(--bad); }}
        .bar-pct {{ font-size: .85rem; color: var(--muted); min-width: 42px; text-align: right; }}
        .heat-na {{ color: var(--muted); }}
        .heat-hi {{ color: var(--ok); font-weight: 600; }}
        .heat-mid {{ color: var(--warn); }}
        .heat-lo {{ color: var(--bad); }}
        .exam-scorecards {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(320px, 1fr)); gap: 16px; }}
        .exam-scorecard {{ background: var(--surface); border: 1px solid var(--border); border-radius: 18px; padding: 18px; box-shadow: var(--shadow); }}
        .exam-scorecard-header {{ display: flex; justify-content: space-between; align-items: baseline; margin-bottom: 14px; gap: 10px; }}
        .exam-scorecard-title {{ font-size: .95rem; font-weight: 700; }}
        .exam-scorecard-meta {{ font-size: .78rem; color: var(--muted); }}
        .scorecard-row {{ display: flex; align-items: center; gap: 8px; margin-bottom: 7px; }}
        .scorecard-model {{ font-size: .8rem; min-width: 120px; max-width: 120px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
        .scorecard-bar-bg {{ flex: 1; height: 7px; background: #efe2d3; border-radius: 4px; overflow: hidden; }}
        .scorecard-bar-fill {{ height: 100%; border-radius: 4px; }}
        .scorecard-pct {{ font-size: .78rem; color: var(--muted); min-width: 38px; text-align: right; }}
        .section-title {{ font-size: .9rem; font-weight: 700; letter-spacing: .05em; text-transform: uppercase; color: var(--muted); margin: 24px 0 12px; }}
        .subject-section {{ margin-top: 48px; scroll-margin-top: calc(var(--nav-h) + 12px); }}
        .subject-header {{ margin-bottom: 16px; padding-bottom: 10px; border-bottom: 2px solid rgba(184,92,56,.2); }}
        .subject-header h2 {{ margin: 0; font-size: clamp(1.5rem, 2vw, 2rem); }}
        .exam-section {{ margin-top: 24px; scroll-margin-top: calc(var(--nav-h) + 12px); }}
        .exam-header {{ display: flex; justify-content: space-between; gap: 12px; align-items: baseline; margin-bottom: 8px; }}
        .exam-header h3 {{ margin: 0; font-size: 1.15rem; }}
        .exam-code {{ color: var(--muted); font-size: .95rem; font-family: 'Courier New', monospace; }}
        .question-card {{ margin-top: 28px; background: rgba(255,253,248,.88); border: 1px solid var(--border); border-radius: 28px; padding: 24px; box-shadow: var(--shadow); backdrop-filter: blur(4px); }}
        .question-topline {{ display: flex; flex-wrap: wrap; gap: 10px; margin-bottom: 12px; }}
        .exam-pill,.session-pill,.question-pill,.type-pill,.points-pill,.badge {{ display: inline-flex; align-items: center; border-radius: 999px; padding: 6px 12px; font-size: .9rem; font-weight: 600; }}
        .exam-pill {{ background: var(--accent); color: white; }}
        .session-pill {{ background: #f5dfc6; color: #6d4c37; }}
        .question-pill {{ background: #f0e3d1; }}
        .type-pill {{ background: #ede7dc; color: #53473f; }}
        .points-pill {{ background: #f7ead8; color: #6d4c37; }}
        .details-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 16px; margin: 18px 0; }}
        .responses-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(320px, 1fr)); gap: 16px; margin-top: 18px; }}
        .response-card {{ background: var(--surface); border: 1px solid #eadbc9; border-radius: 22px; padding: 18px; box-shadow: 0 10px 30px rgba(71,48,26,.05); }}
        .response-header {{ display: flex; justify-content: space-between; gap: 12px; align-items: flex-start; margin-bottom: 10px; }}
        .response-header h3 {{ margin: 0; font-size: 1.1rem; }}
        .badges {{ display: flex; gap: 8px; flex-wrap: wrap; justify-content: flex-end; }}
        .badge {{ background: #ece2d3; color: #3e342e; }}
        .badge-ok {{ background: rgba(47,125,74,.14); color: var(--ok); }}
        .badge-warn {{ background: rgba(157,92,13,.14); color: var(--warn); }}
        .score-high {{ background: rgba(47,125,74,.14); color: var(--ok); }}
        .score-mid {{ background: rgba(157,92,13,.14); color: var(--warn); }}
        .score-low {{ background: rgba(166,63,63,.14); color: var(--bad); }}
        .meta {{ color: var(--muted); font-size: .92rem; margin-bottom: 12px; }}
        .panel {{ background: #fffaf3; border: 1px solid #efe1cf; border-radius: 16px; padding: 14px; margin-top: 12px; }}
        .panel-title {{ font-size: .82rem; letter-spacing: .08em; text-transform: uppercase; color: var(--muted); margin-bottom: 8px; }}
        .panel-body {{ line-height: 1.6; white-space: normal; word-break: break-word; }}
        .context-block {{ margin-top: 10px; }}
        @media (max-width: 720px) {{
            .page {{ padding: 20px 14px 40px; }}
            .question-card, .hero {{ padding: 18px; border-radius: 22px; }}
            .response-header {{ flex-direction: column; }}
            .badges {{ justify-content: flex-start; }}
        }}
    </style>
</head>
<body>
    <nav class="topnav">{nav_links}</nav>
    <main class="page">
        <section class="hero">
            <h1>Matura Models Comparison</h1>
            <p class="subtitle">Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M')} &mdash; {len(all_models)} modeli, {len(all_exams)} egzaminów</p>
        </section>

        <div class="stats-wrap">
          <div class="stats-row">
            <div class="table-box">
              <caption>Ranking modeli</caption>
              <div class="table-scroll">
                <table class="summary-table">
                  <thead><tr><th>#</th><th>Model</th><th>Pytania</th><th>Poprawne</th><th class="bar-cell">Celność</th><th>Śr. wynik</th><th>Koszt $</th></tr></thead>
                  <tbody>{''.join(ranking_rows)}</tbody>
                </table>
              </div>
            </div>
            <div class="table-box">
              <caption>Celność wg typu pytania</caption>
              <div class="table-scroll">
                <table class="summary-table">
                  <thead><tr><th>Typ</th>{model_header_cells}</tr></thead>
                  <tbody>{''.join(type_stats_rows)}</tbody>
                </table>
              </div>
            </div>
          </div>

          <div class="table-box">
            <caption>Celność wg egzaminu</caption>
            <div class="table-scroll">
              <table class="summary-table">
                <thead><tr><th>Egzamin</th>{model_header_cells}</tr></thead>
                <tbody>{''.join(exam_stats_rows_html)}</tbody>
              </table>
            </div>
          </div>

          <div>
            <p class="section-title">Wyniki per egzamin</p>
            <div class="exam-scorecards">
              {scorecards_html}
            </div>
          </div>
        </div>

        {''.join(grouped_sections)}
    </main>
</body>
</html>
"""

        Path(output_path).write_text(html_output, encoding='utf-8')
        print(f"✅ HTML report saved to: {output_path}")


def main():
    """Main evaluation workflow"""

    parser = argparse.ArgumentParser(description='Evaluate model responses and generate reports')
    parser.add_argument(
        '--output', '-o',
        default=None,
        help='Output file stem (without extension). Default: results/comparison_YYYYMMDD_HHMMSS'
    )
    parser.add_argument(
        '--excel',
        action='store_true',
        default=False,
        help='Also generate an Excel report (off by default)'
    )
    args = parser.parse_args()

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    stem = args.output if args.output else f'results/comparison_{timestamp}'

    print("="*70)
    print("STEP 3: EVALUATE RESPONSES")
    print("="*70)
    
    if not os.getenv('OPENROUTER_API_KEY'):
        print("\n❌ OPENROUTER_API_KEY required for evaluation")
        return

    print(f"\n🤖 Evaluator model: {EVALUATOR_MODEL}")
    
    db = QuestionDatabase()
    
    # Get all responses that need evaluation
    conn = db.conn
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT mr.id, mr.question_id, mr.model_name, mr.response,
               q.exam_name, q.question_text, q.correct_answer, q.answer_explanation,
               q.question_type, q.max_points, q.context_text
        FROM model_responses mr
        JOIN questions q ON mr.question_id = q.id
        WHERE mr.id NOT IN (SELECT response_id FROM evaluations)
    """)
    
    responses_to_evaluate = cursor.fetchall()
    
    if not responses_to_evaluate:
        print("\n✅ All responses already evaluated")
    else:
        print(f"\n📝 Evaluating {len(responses_to_evaluate)} responses...")
        total_prompt_tokens = 0
        total_completion_tokens = 0
        
        for resp in tqdm(responses_to_evaluate, desc="Evaluating"):
            (
                resp_id,
                q_id,
                model_name,
                response,
                exam_name,
                q_text,
                correct_ans,
                answer_explanation,
                q_type,
                max_points,
                context_text,
            ) = resp
            
            question = {
                'exam_name': exam_name,
                'question_text': q_text,
                'question_type': q_type,
                'answer_explanation': answer_explanation,
                'max_points': max_points,
                'context_text': context_text,
            }
            
            # Evaluate
            eval_result = evaluate_response_with_llm(question, response, correct_ans)
            
            # Save evaluation
            cursor.execute("""
                INSERT INTO evaluations (response_id, score, is_correct, evaluator_notes)
                VALUES (?, ?, ?, ?)
            """, (
                resp_id,
                eval_result['score'],
                eval_result['is_correct'],
                eval_result['notes']
            ))
            conn.commit()

            total_prompt_tokens += eval_result.get('prompt_tokens', 0)
            total_completion_tokens += eval_result.get('completion_tokens', 0)

        print("\n✅ Evaluation complete")
        if total_prompt_tokens or total_completion_tokens:
            print(
                f"   Tokens used: {total_prompt_tokens} prompt + "
                f"{total_completion_tokens} completion"
            )
    
    # Generate reports
    output_dir = Path(stem).parent
    output_dir.mkdir(parents=True, exist_ok=True)
    html_output_path = Path(f'{stem}.html')

    generate_html_report(db, str(html_output_path))

    if args.excel:
        excel_output_path = Path(f'{stem}.xlsx')
        generate_excel_report(db, str(excel_output_path))
        excel_line = f"\n📊 Excel: {excel_output_path}"
    else:
        excel_line = ''

    print(f"\n{'='*70}")
    print(f"EVALUATION COMPLETE")
    print(f"{'='*70}")
    print(f"\n🌐 HTML:  {html_output_path}{excel_line}")
    
    db.close()


if __name__ == "__main__":
    main()
